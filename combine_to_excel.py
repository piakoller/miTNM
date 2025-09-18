#!/usr/bin/env python3
"""
Combine patient text files and their generated miTNM JSON outputs into a single Excel (.xlsx) summary.

Usage examples (Windows PowerShell):

# Single pair
python .\combine_to_excel.py --patient-file .\patient_example.txt --json-file .\output.json --output-excel .\miTNM_summary.xlsx

# Batch by folder (assumes JSON next to each patient .txt or in --json-dir/./outputs)
python .\combine_to_excel.py --patient-dir .\patients --pattern *.txt --json-dir .\outputs --output-excel .\miTNM_summary.xlsx

Requires: openpyxl
"""
from __future__ import annotations

import argparse
import json
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import List, Optional, Tuple

try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
except Exception as e:
    sys.exit("Error: openpyxl is required to write Excel files. Install with 'pip install openpyxl'.\nDetails: " + str(e))


@dataclass
class Record:
    patient_file: Path
    json_file: Optional[Path]
    miT: str
    miN: str
    miM: str
    confidence: float
    rationale: str
    patient_text: str


def read_text(path: Path) -> str:
    try:
        return path.read_text(encoding="utf-8").strip()
    except FileNotFoundError:
        return ""
    except Exception as e:
        return f"<read error: {e}>"


def parse_json(path: Optional[Path]) -> Tuple[str, str, str, float, str]:
    if not path or not path.exists():
        return ("unknown", "unknown", "unknown", 0.0, "JSON missing")
    try:
        obj = json.loads(path.read_text(encoding="utf-8"))
    except Exception as e:
        return ("unknown", "unknown", "unknown", 0.0, f"JSON parse error: {e}")
    sig = obj.get("miTNM_signature") or {}
    miT = str(sig.get("miT", "unknown") or "unknown").strip()
    miN = str(sig.get("miN", "unknown") or "unknown").strip()
    miM = str(sig.get("miM", "unknown") or "unknown").strip()
    try:
        confidence = float(obj.get("confidence", 0.0))
    except Exception:
        confidence = 0.0
    rationale = str(obj.get("rationale", "")).strip()
    return (miT, miN, miM, max(0.0, min(1.0, confidence)), rationale)


def autosize(ws) -> None:
    for column_cells in ws.columns:
        length = 0
        col = column_cells[0].column
        for cell in column_cells:
            try:
                cell_len = len(str(cell.value)) if cell.value is not None else 0
                length = max(length, cell_len)
            except Exception:
                pass
        ws.column_dimensions[get_column_letter(col)].width = min(80, max(12, length + 2))


def build_records_single(patient_file: Path, json_file: Optional[Path], max_text_chars: int) -> List[Record]:
    text = read_text(patient_file)
    miT, miN, miM, confidence, rationale = parse_json(json_file)
    if max_text_chars >= 0:
        text = (text[:max_text_chars] + ("…" if len(text) > max_text_chars else ""))
    return [Record(patient_file=patient_file, json_file=json_file, miT=miT, miN=miN, miM=miM, confidence=confidence, rationale=rationale, patient_text=text)]


def build_records_batch(patient_dir: Path, pattern: str, json_dir: Optional[Path], default_outputs_dir: Path, max_text_chars: int) -> List[Record]:
    files = sorted(patient_dir.glob(pattern))
    records: List[Record] = []
    for pf in files:
        # Try to find matching JSON by stem
        candidates = []
        candidates.append(pf.with_suffix('.json'))
        if json_dir:
            candidates.append(json_dir / (pf.stem + '.json'))
        # also try default outputs dir if provided
        if default_outputs_dir.exists():
            candidates.append(default_outputs_dir / (pf.stem + '.json'))
        jf = next((c for c in candidates if c and c.exists()), None)
        records.extend(build_records_single(pf, jf, max_text_chars))
    return records


def write_excel(records: List[Record], out_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws.append([
        "patient_file",
        "json_file",
        "miT",
        "miN",
        "miM",
        "confidence",
        "rationale",
        "patient_text",
    ])
    for r in records:
        ws.append([
            str(r.patient_file),
            str(r.json_file) if r.json_file else "",
            r.miT, r.miN, r.miM,
            r.confidence,
            r.rationale,
            r.patient_text,
        ])
    autosize(ws)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))


def parse_args(argv: List[str]) -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Combine patient texts and miTNM JSON outputs into an Excel summary")
    # Single pair
    p.add_argument("--patient-file", help="Path to a single patient .txt file")
    p.add_argument("--json-file", help="Path to the matching JSON output file")
    # Batch mode
    p.add_argument("--patient-dir", help="Directory containing patient .txt files (batch mode)")
    p.add_argument("--pattern", default="*.txt", help="Glob pattern for patient files (default: *.txt)")
    p.add_argument("--json-dir", help="Directory containing JSON outputs (optional)")
    p.add_argument("--output-excel", default="miTNM_summary.xlsx", help="Path to write the Excel summary (.xlsx)")
    p.add_argument("--max-text-chars", type=int, default=500, help="Truncate patient text to this many characters (-1 to include full text)")
    return p.parse_args(argv)


def main(argv: List[str]) -> int:
    args = parse_args(argv)

    out_path = Path(args.output_excel)
    default_outputs_dir = Path("outputs")

    records: List[Record] = []
    if args.patient_file and args.json_file:
        records = build_records_single(Path(args.patient_file), Path(args.json_file), args.max_text_chars)
    elif args.patient_dir:
        pdir = Path(args.patient_dir)
        if not pdir.is_dir():
            sys.exit(f"Error: --patient-dir not found or not a directory: {pdir}")
        jdir = Path(args.json_dir) if args.json_dir else None
        if jdir and not jdir.exists():
            sys.exit(f"Error: --json-dir not found: {jdir}")
        records = build_records_batch(pdir, args.pattern, jdir, default_outputs_dir, args.max_text_chars)
        if not records:
            sys.exit("No patient files found to process.")
    else:
        # Convenience default: if patient_example.txt and output.json exist
        pf = Path("patient_example.txt")
        jf = Path("output.json")
        if pf.exists() and jf.exists():
            records = build_records_single(pf, jf, args.max_text_chars)
        else:
            sys.exit("Provide either --patient-file and --json-file, or --patient-dir.")

    write_excel(records, out_path)
    print(f"Wrote Excel summary: {out_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
